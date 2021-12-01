import configparser
from openpyxl import load_workbook

config = configparser.RawConfigParser()   
configFilePath = r'ini.conf'
config.read(configFilePath)
input_file = config['EXCELFILE']['Path']

wb = load_workbook(filename=input_file)
ws = wb.active

print(ws.max_row - 1)
for row in ws.iter_rows(min_row=2, max_col=3):
    name = row[0].value
    phone = row[1].value
    cccd = row[2].value
    print(name, phone, cccd)